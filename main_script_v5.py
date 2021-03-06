# -*- coding: utf-8 -*-
"""
Created on Fri Nov 20 10:28:43 2020

@author: ovindu wijethunge
"""
from nltk.tokenize import word_tokenize
from func_geting_features import no_sentences, com_length,no_punctuations, count_stop_word,check_black_words_list,duplicate_words, find_mail_site_pnumber,calculate_content_comment_similerity
from func_vector_create import vectorize
from func_preprocessing import preprocess,remove_punctuation
from func_find_duplicate_comments import find_duplicate_comments
from func_make_csv import makes_csv, create_file
import numpy as np
from datetime import datetime
import openpyxl
import re

# client id = 797659646664-sc1saj14036j4ulm5u49bimpsm1cgitk.apps.googleusercontent.com
# client secret =  NaEFFaU2WXbAnA_jynaH7t26

def mainScript():
    
    wb1 = openpyxl.load_workbook('contentData.xlsx')
    wb2 = openpyxl.load_workbook('commentData.xlsx') 
    
    worksheet1 = wb1.active
    worksheet2 = wb2.active
    number_of_round = (worksheet1.max_row + 1)
    fields_list = []
    for x in range(2,number_of_round):
       # print("round????????????????????????????????????????????????????????????????????????????????????????? ")
        vid = worksheet1.cell(row=x, column=2).value
        main_list = []
        post_date = ""
        for row in worksheet2.iter_rows():
            
            if row[0].value == vid:
                row_list = []
                count = 1
                for cell in row:
                    row_list.append(worksheet2.cell(row=cell.row, column=count).value)
                    post_date = worksheet2.cell(row=cell.row, column=2).value
                    count +=1
                main_list.append(row_list)
                
        print('your lis ',main_list)
        dataset = []
        comments_date = []
        comment_id = []
        for li in main_list:
            dataset.append(li[3])
            comments_date.append(li[8])
            comment_id.append(li[2])
        content_string = worksheet1.cell(row=x, column=1).value
        dataset.insert(0,content_string)
        N = len(dataset)
        
        
        def calculate_time_gap(val):
            rep1 = post_date.replace("Z", "+00:00")
            rep2 = val.replace("Z", "+00:00")
            
            x = datetime.fromisoformat(rep1).timestamp()
            y = datetime.fromisoformat(rep2).timestamp()
            dif = (y-x)/60 # calculate from minitues
            return dif
        
        def generate_ngrams(text):
            
            text = re.sub(r'[#$%&()*+-./:;<=>?\s]', '',text)
            tokens = word_tokenize(text)
            print(tokens) # ['මචං', 'මැසේජ්', 'වගේ']
            list_comment = []
            for token in tokens:# take a one word at  a time
                n = len(token)
                list_word = []
                for i in range(0,(n-1)):
                    bi = token[i:i+2]
                    list_word.append(bi)
                print("word bi grams list ",list_word)    
                list_comment.extend(list_word)
        
            print("comment bi grams ",list_comment)
            return list_comment
        
        
        
        
    
        bi_gram_processed_text = [] #this is a list in list which have bi-gram strings duplicate bi-grams also here
        words_processed_text = [] # this is a list in list which have words tokens strings duplicate words also here
        
        
        for text in dataset[:N]:
            text_preprocessed_bi_gram = preprocess(text) # remove stop words and punctuations for make bi grams for find similerity
            bi_grams = generate_ngrams(text_preprocessed_bi_gram)# make a character bigram list for a comment
            bi_gram_processed_text.append(bi_grams)# make a list in list of bigrams for each comments
            
            
            words_processed_text.append(word_tokenize(str(remove_punctuation(text))))
            
        vectors_list = vectorize(bi_gram_processed_text)
    
    #---------------------------------------------------------------------------------------
    
        center_vector = np.zeros((len(vectors_list[2])))
        for numb in range(1,len(vectors_list)):
            center_vector = center_vector + vectors_list[numb]
        new_center_vector = center_vector/(len(vectors_list)-1)        
        content_vector = vectors_list[0]
        
        for i in range(1, N):
            cos_simillerity_content_comment = calculate_content_comment_similerity(content_vector,vectors_list[i].tolist())
            cos_simillerity_comment_comment = calculate_content_comment_similerity(new_center_vector,vectors_list[i].tolist())    
            word_count =len(words_processed_text[i]) 
            duplicate_word_ratio = duplicate_words(words_processed_text[i])
            no_of_sentences = no_sentences(dataset[i])
            length_of_comment = com_length(dataset[i])
            num_of_punctuations = no_punctuations(dataset[i])
            stop_word_ratio = count_stop_word(words_processed_text[i])
            post_coment_gap = calculate_time_gap(comments_date[i-1])
            is_black_word = check_black_words_list(words_processed_text[i])
            link_mail_pnumber = find_mail_site_pnumber(dataset[i])
            comment_duplication = find_duplicate_comments(main_list , i-1)
            
            comment_id_val = comment_id[i-1]
# =============================================================================
#             print(dataset[i])
#             print("cos_simillerity_content_comment ",cos_simillerity_content_comment)
#             print("cos_simillerity_comment_comment  ",cos_simillerity_comment_comment)
#             print("word_count  ",word_count)
#             print("duplicate_word_ratio  ",duplicate_word_ratio)
#             print("no_of_sentences  ",no_of_sentences)
#             print("length_of_comment ",length_of_comment)
#             print("num_of_punctuations ",num_of_punctuations)
#             print("stop_word_ratio ",stop_word_ratio)
#             print("post_coment_gap ",post_coment_gap)
#             print("is_black_word ",is_black_word)
#             print("link_mail_pnumber ",link_mail_pnumber)
#             print("comment_duplication  ",comment_duplication)
# =============================================================================
 #   ['BHhl75a5bks', '2021-01-29T14:34:46Z', 'Ugw5ej-XdiFL40uu9814AaABAg', 'හැමදාම unea ඔහොම තමා', 'UCh3O7jnH1dspTmxbkW4K6yA', 'ශ්\u200dරී ලංකා ක්\u200dරිකට් තවත් කැපිල්ලක්..', 0, '2021-01-30T16:12:28Z', '2021-01-30T16:12:28Z']
           
            stringList = makes_csv(comment_id_val, cos_simillerity_content_comment,cos_simillerity_comment_comment,word_count,duplicate_word_ratio,no_of_sentences,length_of_comment,num_of_punctuations,stop_word_ratio,post_coment_gap,is_black_word,link_mail_pnumber,comment_duplication)
            fields_list.append(stringList)
            
            
        create_file(fields_list)       

#mainScript()    